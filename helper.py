import os
import numpy as np
import pandas as pd
import boto3
import time
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def read_s3(bucket_name, prefix, keywords):
    s3 = boto3.client('s3')
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

    if 'Contents' in response:
        files = [obj['Key'] for obj in response['Contents'] if obj['Key'].endswith(keywords)]
    else:
        print("No files found.")

    # next is to read them all into pandas dataframes
    df4 = []
    for file in files:
        path = 's3://'+ bucket_name + '/' + file
        print('Reading from the file ', file)
        df = pd.read_csv(path, sep = ',', low_memory=False)
        df4.append(df)


    # final step is to catch into one dataframe
    if df4:
        results = pd.concat(df4, ignore_index=True)
        print('Load file successfully, file length is ', len(results))
        print('Now the total rows are ', len(results))
    else:
        results = pd.DataFrame()
        print("No files found.")
    
    results.columns = results.columns.str.lower()
    return results





def modify_excel(out_path):
    wb = load_workbook(out_path)

    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue 
        ws = wb[sheet_name]
        ws.insert_rows(1)

        # ws['C1'] = 'The data from V4'
        ws.insert_cols(1)
        # ws['A2'] = 'The data from V5'
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value))
                    max_len = max(max_len, val_len)
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_len * 1.2

    wb.save(out_path)
    print('The Excel is successfully modified')


def keywords_in_var(keywords, var):
    if not isinstance(keywords, str):
        for keyword in keywords:
            key_columns = [col for col in var.columns if keyword in col.lower()]
            print('There are ', len(key_columns), ' variables contains ' + keyword + '.')
            print(" ".join(key_columns))
    else:
        print('keywords are not in iterable')
        key_columns = [col for col in var.columns if keywords in col.lower()]
        print('There are ', len(key_columns), ' variables contains ' + keywords + '.')
        print(" \n".join(key_columns))


_INFLATION_DATA = [
    (2005, 1.034, 2.453), (2006, 1.032, 2.376), (2007, 1.029, 2.310),
    (2008, 1.038, 2.225), (2009, 1.020, 2.181), (2010, 1.020, 2.138),
    (2011, 1.031, 2.073), (2012, 1.021, 2.031), (2013, 1.021, 1.991),
    (2014, 1.020, 1.952), (2015, 1.019, 1.914), (2016, 1.020, 1.877),
    (2017, 1.021, 1.837), (2018, 1.024, 1.794), (2019, 1.020, 1.758),
    (2020, 1.021, 1.724), (2021, 1.188, 1.452), (2022, 1.168, 1.243),
    (2023, 1.086, 1.145), (2024, 1.070, 1.070), (2025, 1.070, 1.000),
]
_INFLATION_DF = pd.DataFrame(
    _INFLATION_DATA, columns=["year", "inflation_factor", "cumulative_inflation_factor"]
).astype({"year": "Int64"})

""" 
_INFLATION_DATA = [
    (2005, 2.171),
    (2006, 2.103),
    (2007, 2.044),
    (2008, 1.969),
    (2009, 1.931),
    (2010, 1.893),
    (2011, 1.835),
    (2012, 1.798),
    (2013, 1.763),
    (2014, 1.728),
    (2015, 1.694),
    (2016, 1.661),
    (2017, 1.626),
    (2018, 1.588),
    (2019, 1.556),
    (2020, 1.526),
    (2021, 1.285),
    (2022, 1.1),
    (2023, 1)
]
_INFLATION_DF = pd.DataFrame(
    _INFLATION_DATA, columns=["year", "cumulative_inflation_factor"]
).astype({"year": "Int64"})

 """
def add_cumulative_inflation(
    df: pd.DataFrame,
    *,
    year_col: str = "year",
    loss_col: str = "ncat",
    ee_col: str = "ee",
    compute_pp: bool = False,
) -> pd.DataFrame:
    """
    Merge cumulative inflation by year and add 'ncat_infl_adj'.
    Optionally compute 'pp_infl_adj' = ncat_infl_adj / ee (skip when ee==0).

    Parameters
    ----------
    df : input dataframe with `year_col` and `loss_col`.
    year_col : name of year column in df.
    loss_col : name of loss column to inflate (default 'ncat').
    ee_col : exposure column for pp (used only if compute_pp=True).
    compute_pp : if True, also adds 'pp_infl_adj'.
    inflation_df : optional custom inflation table with columns ['year','cumulative_inflation_factor'].

    Returns
    -------
    pd.DataFrame : a copy of df with added columns:
        - cumulative_inflation_factor
        - ncat_infl_adj
        - (optional) pp_infl_adj
    """
    if year_col not in df.columns:
        raise KeyError(f"Missing year column {year_col!r}")
    if loss_col not in df.columns:
        raise KeyError(f"Missing loss column {loss_col!r}")

    infl = _INFLATION_DF.copy()
    infl["year"] = pd.to_numeric(infl["year"], errors="coerce").astype("Int64")

    d = df.copy()
    d[year_col] = pd.to_numeric(d[year_col], errors="coerce").astype("Int64")

    # Merge on the user's year column name
    d = d.merge(
        infl.rename(columns={"year": year_col})[[year_col, "cumulative_inflation_factor"]],
        on=year_col, how="left"
    )

    d["cumulative_inflation_factor"] = d["cumulative_inflation_factor"].fillna(1.0)
    d["ncat_infl_adj"] = d[loss_col] * d["cumulative_inflation_factor"]

    if compute_pp and ee_col in d.columns:
        d["pp_infl_adj"] = d["ncat_infl_adj"] / d[ee_col].replace({0: np.nan})
    
    print('The inflation adjusted ncat is added into the dataframe as ncat_infl_adj.')
    if compute_pp:
        print('The inflation adjusted pp is added into the dataframe as pp_infl_adj.')

    return d

def format_number(n):
    if abs(n) >= 1_000_000_000:
        return f'{n / 1_000_000_000:.2f}B'
    elif abs(n) >= 1_000_000:
        return f'{n / 1_000_000:.2f}M'
    elif abs(n) >= 1_000:
        return f'{n / 1_000:.2f}K'
    else:
        return str(round(n, 2))
